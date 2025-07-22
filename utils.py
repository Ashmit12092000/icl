import io
import logging
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter # <--- Added this import
from sqlalchemy import and_
# Configure logging for utils.py
logging.basicConfig(level=logging.DEBUG)

def compute_cumulative_principal(transactions, new_paid=Decimal('0'), new_repaid=Decimal('0')):
    """
    Calculates cumulative principal for the next compounding period, including all prior net interest.
    - transactions: list of Transaction objects, sorted by date, each must have .get_safe_amount_paid(), .get_safe_amount_repaid(), and .get_safe_net_amount()
    - new_paid: Decimal, amount_paid for the new transaction (if any, else 0)
    - new_repaid: Decimal, amount_repaid for the new transaction (if any, else 0)
    Returns: Decimal, principal to use for interest calculation
    """
    cumulative = Decimal('0')
    for txn in transactions:
        cumulative += txn.get_safe_amount_paid()
        cumulative -= txn.get_safe_amount_repaid()
        cumulative += txn.get_safe_net_amount()  # Net interest is compounded
    cumulative += new_paid
    cumulative -= new_repaid
    return cumulative
def safe_decimal_conversion(value, default=Decimal('0')):
    """Safely convert values to Decimal, handling None and invalid inputs"""
    if value is None or value == '':
        return default
    
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return default

def calculate_interest(principal, annual_rate, days):
    """Calculate simple interest for given principal, annual rate, and number of days"""
    if principal <= 0 or annual_rate <= 0 or days <= 0:
        return Decimal('0')
    
    # Simple interest formula: (Principal * Rate * Time) / 100
    # Time is in years, so days/365
    interest = (principal * annual_rate * Decimal(str(days))) / (Decimal('100') * Decimal('365'))
    return interest.quantize(Decimal('0.01'))

def calculate_compound_interest(principal, annual_rate, days, compounding_frequency='quarterly'):
    """Calculate compound interest for given principal, annual rate, days, and compounding frequency"""
    if principal <= 0 or annual_rate <= 0 or days <= 0:
        return Decimal('0')
    
    # For now, use simple interest calculation as compound logic is handled in transaction processing
    return calculate_interest(principal, annual_rate, days)

def export_to_excel(customer, transactions):
    """Export customer and transaction data to Excel format"""
    output = io.BytesIO()
    
    # Create workbook and worksheet
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = f"Customer Report - {customer.icl_no}"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center')
    
    # Customer information section
    worksheet['A1'] = "Customer Information"
    worksheet['A1'].font = header_font
    worksheet['A1'].fill = header_fill
    worksheet.merge_cells('A1:B1')
    
    customer_info = [
        ["ICL No:", customer.icl_no],
        ["Name:", customer.name],
        ["Address:", customer.address],
        ["Contact:", customer.contact_details],
        ["Annual Rate:", f"{customer.annual_rate}%"],
        ["ICL Start Date:", customer.icl_start_date.strftime('%d-%m-%Y') if customer.icl_start_date else ""],
        ["ICL End Date:", customer.icl_end_date.strftime('%d-%m-%Y') if customer.icl_end_date else ""],
        ["Interest Type:", customer.interest_type],
        ["TDS Applicable:", "Yes" if customer.tds_applicable else "No"],
        ["TDS Percentage:", f"{customer.tds_percentage}%" if customer.tds_percentage else "0%"]
    ]
    
    for i, (label, value) in enumerate(customer_info, start=2):
        worksheet[f'A{i}'] = label
        worksheet[f'B{i}'] = value
        worksheet[f'A{i}'].border = border
        worksheet[f'B{i}'].border = border
    
    # Transaction header
    start_row = len(customer_info) + 4
    headers = [
        "Date", "Amount Paid", "Amount Repaid", "Balance", 
        "Period From", "Period To", "Days", "Interest Rate", 
        "Interest Amount", "TDS Amount", "Net Amount"
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=start_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment
    
    # Process transactions and consolidate ICL end date repayments
    processed_transactions = []
    i = 0
    
    while i < len(transactions):
        transaction = transactions[i]
        
        # Check if this is an ICL end date scenario with multiple transactions
        is_icl_end_date = (
            customer.icl_end_date and 
            transaction.date == customer.icl_end_date
        )
        
        if is_icl_end_date:
            # Collect all transactions on the ICL end date
            icl_end_transactions = [transaction]
            j = i + 1
            
            while j < len(transactions) and transactions[j].date == customer.icl_end_date:
                icl_end_transactions.append(transactions[j])
                j += 1
            
            # Separate different types of transactions
            period_txn = None
            repayment_txn = None
            loan_closure_txn = None
            
            for txn in icl_end_transactions:
                if txn.transaction_type == 'repayment' and txn.amount_repaid and txn.amount_repaid > Decimal('0'):
                    repayment_txn = txn
                elif txn.transaction_type == 'loan_closure':
                    loan_closure_txn = txn
                elif txn.period_from and txn.period_to and txn.int_amount:
                    # This is the period interest calculation transaction
                    period_txn = txn
            
            # For ICL end date, create a single consolidated entry that shows:
            # 1. The period with interest calculation
            # 2. The repayment amount in the same row
            if repayment_txn and repayment_txn.period_from and repayment_txn.period_to:
                # Create single consolidated entry showing both period interest and repayment
                consolidated_data = {
                    'date': repayment_txn.period_from,  # Show period start date
                    'period_from': repayment_txn.period_from,
                    'period_to': repayment_txn.period_to,
                    'no_of_days': repayment_txn.no_of_days,
                    'int_rate': repayment_txn.int_rate,
                    'int_amount': repayment_txn.int_amount,
                    'tds_amount': repayment_txn.tds_amount,
                    'net_amount': repayment_txn.net_amount,
                    'amount_paid': repayment_txn.amount_paid,
                    'amount_repaid': None,  # Don't show repayment in period row
                    'balance': repayment_txn.balance
                }
                processed_transactions.append(consolidated_data)
                
                # Create separate repayment row (simple repayment entry)
                repayment_data = {
                    'date': repayment_txn.date,
                    'period_from': None,
                    'period_to': None,
                    'no_of_days': None,
                    'int_rate': None,
                    'int_amount': None,
                    'tds_amount': None,
                    'net_amount': None,
                    'amount_paid': None,
                    'amount_repaid': repayment_txn.amount_repaid,
                    'balance': Decimal('0')  # Always zero after full repayment
                }
                processed_transactions.append(repayment_data)
            else:
                # Fallback: add each transaction individually if consolidation fails
                for txn in icl_end_transactions:
                    if txn.transaction_type != 'loan_closure':  # Skip loan closure entries
                        transaction_data = {
                            'date': txn.date,
                            'period_from': txn.period_from,
                            'period_to': txn.period_to,
                            'no_of_days': txn.no_of_days,
                            'int_rate': txn.int_rate,
                            'int_amount': txn.int_amount,
                            'tds_amount': txn.tds_amount,
                            'net_amount': txn.net_amount,
                            'amount_paid': txn.amount_paid,
                            'amount_repaid': txn.amount_repaid,
                            'balance': txn.balance
                        }
                        processed_transactions.append(transaction_data)
            
            # Skip all processed transactions
            i = j
        else:
            # Regular transaction - add as is
            transaction_data = {
                'date': transaction.date,
                'period_from': transaction.period_from,
                'period_to': transaction.period_to,
                'no_of_days': transaction.no_of_days,
                'int_rate': transaction.int_rate,
                'int_amount': transaction.int_amount,
                'tds_amount': transaction.tds_amount,
                'net_amount': transaction.net_amount,
                'amount_paid': transaction.amount_paid,
                'amount_repaid': transaction.amount_repaid,
                'balance': transaction.balance
            }
            processed_transactions.append(transaction_data)
            i += 1
    
    # Transaction data
    total_days = Decimal('0')
    total_interest = Decimal('0')
    total_tds = Decimal('0')
    total_net_amount = Decimal('0')
    
    for row, txn_data in enumerate(processed_transactions, start=start_row + 1):
        # Accumulate totals
        if txn_data['no_of_days']:
            total_days += Decimal(str(txn_data['no_of_days']))
        if txn_data['int_amount']:
            total_interest += txn_data['int_amount']
        if txn_data['tds_amount']:
            total_tds += txn_data['tds_amount']
        if txn_data['net_amount']:
            total_net_amount += txn_data['net_amount']
        
        data = [
            txn_data['date'].strftime('%d-%m-%Y') if txn_data['date'] else "",
            str(txn_data['amount_paid']) if txn_data['amount_paid'] else "",
            str(txn_data['amount_repaid']) if txn_data['amount_repaid'] else "",
            str(txn_data['balance']) if txn_data['balance'] is not None else "",
            txn_data['period_from'].strftime('%d-%m-%Y') if txn_data['period_from'] else "",
            txn_data['period_to'].strftime('%d-%m-%Y') if txn_data['period_to'] else "",
            str(txn_data['no_of_days']) if txn_data['no_of_days'] else "",
            f"{txn_data['int_rate']}%" if txn_data['int_rate'] else "",
            str(txn_data['int_amount']) if txn_data['int_amount'] else "",
            str(txn_data['tds_amount']) if txn_data['tds_amount'] else "",
            str(txn_data['net_amount']) if txn_data['net_amount'] else ""
        ]
        
        for col, value in enumerate(data, start=1):
            cell = worksheet.cell(row=row, column=col, value=value)
            cell.border = border
            if col > 1:  # Align numbers to right
                cell.alignment = Alignment(horizontal='right')
    
    # Add totals row if there are processed transactions
    if processed_transactions:
        totals_row = len(processed_transactions) + start_row + 2  # Add space between data and totals
        
        # Add "TOTAL" label
        total_label_cell = worksheet.cell(row=totals_row, column=6, value="TOTAL:")
        total_label_cell.font = Font(bold=True)
        total_label_cell.border = border
        total_label_cell.alignment = Alignment(horizontal='right')
        
        # Add totals for specific columns
        totals_data = [
            str(int(total_days)),  # Column 7: Total Days
            "",  # Column 8: Interest Rate (no total needed)
            str(total_interest.quantize(Decimal('0.01'))),  # Column 9: Total Interest Amount
            str(total_tds.quantize(Decimal('0.01'))),  # Column 10: Total TDS Amount
            str(total_net_amount.quantize(Decimal('0.01')))  # Column 11: Total Net Amount
        ]
        
        for col, value in enumerate(totals_data, start=7):
            cell = worksheet.cell(row=totals_row, column=col, value=value)
            cell.font = Font(bold=True)
            cell.border = border
            cell.alignment = Alignment(horizontal='right')
            if col > 8:  # Add background color to totals columns
                cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    workbook.save(output)
    output.seek(0)
    return output

def get_period_report(start_date, end_date):
    """Generate period report for all customers within date range"""
    from models import Customer, Transaction
    
    output = io.BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = f"Period Report {start_date} to {end_date}"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        "ICL No", "Customer Name", "Total Paid", "Total Repaid", 
        "Total Interest", "Total TDS", "Net Interest", "Current Balance"
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    # Get all active customers
    customers = Customer.query.filter_by(is_active=True).all()
    
    row = 2
    for customer in customers:
        # Get transactions in the period
        transactions = Transaction.query.filter(
            and_(
                Transaction.customer_id == customer.id,
                Transaction.date >= start_date,
                Transaction.date <= end_date
            )
        ).all()
        
        if transactions:  # Only include customers with transactions in the period
            total_paid = sum(t.get_safe_amount_paid() for t in transactions)
            total_repaid = sum(t.get_safe_amount_repaid() for t in transactions)
            total_interest = sum(t.get_safe_int_amount() for t in transactions)
            total_tds = sum(t.get_safe_tds_amount() for t in transactions)
            total_net_interest = sum(t.get_safe_net_amount() for t in transactions)
            current_balance = customer.get_current_balance()
            
            data = [
                customer.icl_no,
                customer.name,
                str(total_paid),
                str(total_repaid),
                str(total_interest),
                str(total_tds),
                str(total_net_interest),
                str(current_balance)
            ]
            
            for col, value in enumerate(data, start=1):
                cell = worksheet.cell(row=row, column=col, value=value)
                cell.border = border
            
            row += 1
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    workbook.save(output)
    output.seek(0)
    return output
